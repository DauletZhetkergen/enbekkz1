import configparser
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import date
import os.path
config = configparser.ConfigParser()
config.read("settings.ini")

webDriverPath = config["BASIC"]["WEBDRIVER"]
admin = config["ADMIN"]["USER"]
password = config["ADMIN"]["PASS"]


nameExcel = ""


def main(roles=0,name="",groupNumber = None,group=False):
    global nameExcel
    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)
    chrome_options.add_argument("--lang=ru")
    services = Service(rf"C:\Users\abdinur\chromedriver.exe")
    driver = webdriver.Chrome(service=services,options=chrome_options)
    driver.get("https://login.kundelik.kz/login")
    time.sleep(2)
    driver.maximize_window()
    time.sleep(1)


    if group:
        nameExcel=  "Массовая "
    elif roles==1:
        nameExcel += "Ученик"
    elif roles==2:
        nameExcel += "Родитель"
    elif roles==3:
        nameExcel += "УченикиРодитель"
    else:
        nameExcel = "Прочее"

    createExcel(nameExcel)

    # driver.find_element(By.XPATH,"//div[@class='localization-select']//select").click()
    # time.sleep(0.3)
    # slct = Select(driver.find_element(By.XPATH,"//div[@class='localization-select']//select"))
    # slct.select_by_value("ru-RU")
    # driver.find_element(By.XPATH, "//div[@class='localization-select']//select/option[2]").click()

    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@value='Кіру']")))

    #login to site by admin
    driver.find_element(By.XPATH, "//input[@name='login']").send_keys(admin)
    driver.find_element(By.XPATH, "//input[@name='password']").send_keys(password)
    time.sleep(0.5)
    driver.find_element(By.XPATH, "//input[@value='Кіру']").click()
    ###



    #Edu->mySchool->People
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//div[@data-test-id = 'new-post-button']")))


    # driver.find_element(By.XPATH, "//div[contains(@class,'page-header__panel-container')]//div[contains(@class,'header-localization-select__info_row')][1]").click()
    # time.sleep(0.5)
    # driver.find_element(By.XPATH, "//div[contains(@class,'page-header__panel-container')]//a[text()=' Русский ']").click()
    # WebDriverWait(driver, 10).until(
        # EC.visibility_of_element_located((By.XPATH, "//div[@data-test-id = 'new-post-button']")))
    driver.find_element(By.XPATH, "//li/a[@title='Моя школа']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//div/button[@data-test-id='send-post']")))

    driver.find_element(By.XPATH, "//div[@class='tabs']//a[text()='Люди']").click()
    #
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@value='Найти']")))
    #Main process()
    # try:
    if group:
        gotoAllStudent(driver, True)
    elif len(name)>1:
        driver.find_element(By.XPATH, "//div/input[@id='search']").send_keys(name)
        time.sleep(0.5)
        if groupNumber != None:
            driver.find_element(By.XPATH, "//input[@id='class']").send_keys(groupNumber)
        time.sleep(0.5)
        driver.find_element(By.XPATH, "//input[@value='Найти']").click()
        time.sleep(1.5)
        foundStudent = driver.find_element(By.XPATH, "//p[@class='found']/strong").text
        if int(foundStudent) == 1:
            time.sleep(0.5)
            if driver.find_element(By.XPATH,"//p[@class='found']/strong"):
                if roles == 1:
                    goToStudent(driver)
                elif roles == 2:
                    gotoDirectParent(driver)
                elif roles == 3:
                    goToStudent(driver,True)
                elif roles == 4:
                    gotoTeacher(driver)
                else:
                    pass
        else:
            if roles == 1:
                print("kirdi")
                gotoAllStudent(driver, False)
            elif roles == 2:
                gotoAllStudentParent(driver)
            elif roles == 3:
                gotoAllStudent(driver, True)
    elif len(name)==0 and groupNumber != None:
        driver.find_element(By.XPATH, "//input[@id='class']").send_keys(groupNumber)
        driver.find_element(By.XPATH, "//input[@value='Найти']").click()
        if roles == 1:
            gotoAllStudent(driver, False)
        elif roles == 2:
            gotoAllStudentParent(driver)
        elif roles == 3:
            gotoAllStudent(driver, True)
        else:
            gotoAllStudent(driver, False)
    elif roles==1:
        driver.find_element(By.XPATH, "//li[@class='iGroup']/a[text()='Ученики']").click()
        gotoAllStudent(driver, False)
    elif roles ==2:
        driver.find_element(By.XPATH, "//li[@class='iGroup']/a[text()='Родители']").click()
        gotoAllStudentParent(driver)
    elif roles == 3:
        driver.find_element(By.XPATH, "//li[@class='iGroup']/a[text()='Ученики']").click()
        gotoAllStudent(driver,True)
    elif roles == 5:
        driver.find_element(By.XPATH, "//li[@class='iGroup']/a[text()='Сотрудники']").click()
        gotoAllTeacher(driver)
    print(1)
    return True
    # except:
    #     print(2)
    #     return False





def gotoTeacher(driver):
    driver.find_element(By.XPATH,
                        "//table[contains(@class,'people')]/tbody//td[@class='tdButtons']//li[@class='iE']").click()
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']")))

    driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='change']")))

    studLogin = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text

    parent = checkParentsExist(studLogin)
    if parent:
        writeOnlyParent(parent[0], parent[1], parent[2], parent[3])
    else:
        driver.find_element(By.XPATH, "//input[@name='change']").click()
        fullName = driver.find_element(By.XPATH, "//div[@class= 'owner people_page' ]/div//h2").text
        tempPass = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[2]").text

        writeOnlyParent(fullName, studLogin, tempPass, "Учитель")

def gotoAllTeacher(driver):

    time.sleep(0.5)
    lastPage = 1
    try:
        lastPage = driver.find_element(By.XPATH, "//div[@class='pager']//li[last()]/a").text
    except:
        pass
    i = 1
    while i <= int(lastPage):
        numberStudents = len(driver.find_elements(By.XPATH, "//table[@class='people grid']/tbody/tr"))

        j = 1
        while j <= int(numberStudents):
            time.sleep(5)
            typeOfAgent = driver.find_element(By.XPATH,
                                              "//table[@class='people grid']/tbody/tr[{}]//td[@class='tdName']/p".format(
                                                  j)).text

            driver.find_element(By.XPATH,
                                f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdButtons']//li[@class='iE']").click()
            time.sleep(1)
            massiveParent(driver, 'Учитель')
            j += 1
        time.sleep(3)
        i += 1
        driver.find_element(By.XPATH, f"//div[@class='pager']//li/a[text()={i}]").click()
        time.sleep(3)

def gotoAllStudentParent(driver):


    lastPage = 1
    try:
        lastPage = driver.find_element(By.XPATH, "//div[@class='pager']//li[last()]/a").text
    except:
        pass
    i=1
    while i<=int(lastPage):
        numberStudents = len(driver.find_elements(By.XPATH, "//table[@class='people grid']/tbody/tr"))

        j=1
        while j<=int(numberStudents):
            time.sleep(5)
            typeOfAgent = driver.find_element(By.XPATH, "//table[@class='people grid']/tbody/tr[{}]//td[@class='tdName']/p".format(j)).text
            driver.find_element(By.XPATH, f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdButtons']//li[@class='iE']").click()
            time.sleep(1)
            massiveParent(driver,'Родитель')
            j+=1
        time.sleep(3)
        i+=1
        driver.find_element(By.XPATH, f"//div[@class='pager']//li/a[text()={i}]").click()
        time.sleep(3)

def gotoAllStudent(driver,parent):
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//p[@class='found']")))
    lastPage = 1
    try:
        lastPage = driver.find_element(By.XPATH, "//div[@class='pager']//li[last()]/a").text
    except:
        pass
    i=1
    while i<=int(lastPage):
        numberStudents = len(driver.find_elements(By.XPATH, "//table[@class='people grid']/tbody/tr"))

        j=1
        while j<=int(numberStudents):
            try:
                time.sleep(1)
                typeOfAgent = driver.find_element(By.XPATH, "//table[@class='people grid']/tbody/tr[{}]//td[@class='tdName']/p".format(j)).text
                print(typeOfAgent)
                nameOfParent = driver.find_element(By.XPATH,
                                                   f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdName']/a").text
                time.sleep(1)
                if typeOfAgent == 'Ученик':
                    driver.find_element(By.XPATH,
                                        f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdButtons']//li[@class='iE']").click()
                    massiveStud(driver,parent)
                elif "Родитель" in typeOfAgent:
                    statusExist = checkParentExistFullName(nameOfParent)
                    if statusExist:
                        pass
                    else:
                        driver.find_element(By.XPATH,
                                            f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdButtons']//li[@class='iE']").click()
                        massiveParent(driver,'Родитель')
                elif "Сотрудник" in typeOfAgent:
                    print('sotrudnik')
                    driver.find_element(By.XPATH,
                                        f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdButtons']//li[@class='iE']").click()
                    massiveParent(driver,'Сотрудник')
                elif "Учитель" in typeOfAgent:
                    driver.find_element(By.XPATH,
                                        f"//table[@class='people grid']/tbody/tr[{j}]//td[@class='tdButtons']//li[@class='iE']").click()
                    massiveParent(driver, 'Учитель')
                else:
                    pass
                j+=1
            except:
                j += 1
                pass

        if i == int(lastPage):
            print(1)
            return True
        else:
            i += 1
            driver.find_element(By.XPATH, f"//div[@class='pager']//li/a[text()={i}]").click()
            time.sleep(3)

def checkParentExistFullName(nameParent):
    if os.path.isfile(f"{date.today()}_Отчет о выгрузке{nameExcel}.xlsx"):
        wb = openpyxl.load_workbook(f"{date.today()}_Отчет о выгрузке{nameExcel}.xlsx")
    else:
        wb = openpyxl.Workbook()

    sheet = wb.worksheets[0]
    foundParent = 0
    for row in sheet.iter_rows(min_col=6, max_col=6):
        for cell in row:
            if str(cell.value) == str(nameParent):
                foundParent = cell
    if foundParent!=0:
        return True
    return False

def massiveStud(driver,parent):
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//a[@id='TabPassword']")))
    driver.find_element(By.XPATH, f"//a[@id='TabPassword']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='change']")))
    driver.find_element(By.XPATH, "//input[@name='change']").click()

    fullName = driver.find_element(By.XPATH, "//div[@class= 'owner people_page' ]/div//h2").text
    studLogin = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text
    tempPass = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[2]").text
    driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabReview']").click()
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//div[@class='panel blue2'][2]//dl[@class='info big']/dd[1]//a")))
    classNumber = driver.find_element(By.XPATH, "//div[@class='panel blue2'][2]//dl[@class='info big']/dd[1]//a").text

    # Parents part
    driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabParents']").click()

    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//li/a[@id='buttonAddRelative']")))
    parents = None
    if parent:
        try:
            driver.find_element(By.XPATH, "//div[@class='emptyData']")
            driver.back()
            time.sleep(1)
            driver.back()
            time.sleep(1)
            driver.back()
            time.sleep(1)
            driver.back()
            time.sleep(1)
            driver.back()
            time.sleep(1)
        except:
            status = False
            logins = getOnlyLoginParent(driver)
            if logins !=0:
                status = checkParentsExist(logins)
                print(status)
            if status:
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                driver.back()
                a = []
                a.append(status)
                create_write_excel(fullName, studLogin, tempPass, classNumber, a)
                return
            else:
                # driver.back()
                # time.sleep(1)
                # driver.back()
                # time.sleep(1)
                # driver.back()
                # time.sleep(1)
                # driver.back()
                # time.sleep(1)
                # driver.back()
                parents = goToParent(driver)

    else:
        driver.back()
        time.sleep(1)
        driver.back()
        time.sleep(1)
        driver.back()
        time.sleep(1)
        driver.back()
        time.sleep(1)
        driver.back()
    create_write_excel(fullName, studLogin, tempPass, classNumber, parents)

def getOnlyLoginParent(driver):
    rows = driver.find_elements(By.XPATH, "//table[contains(@class,'people')]/tbody/tr")
    numberOfRows = len(rows)
    logins = []
    driver.find_element(By.XPATH,
                        f"//table[contains(@class,'people')]/tbody/tr[2]/td[@class='tdButtons']//li[@class='iE']").click()
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']")))
    driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='change']")))
    parentLogin = driver.find_element(By.XPATH,
                                      "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text
    logins = parentLogin
    driver.back()
    time.sleep(1)
    driver.back()
    time.sleep(1)
    if len(logins)>0:
        return logins
    return 0

def massiveParent(driver,user):
    driver.find_element(By.XPATH, f"//a[@id='TabPassword']").click()

    studLogin = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text
    if studLogin=='admin.sko':
        time.sleep(1)
        driver.back()
        time.sleep(1)
        driver.back()
    else:
        parent = checkParentsExist(studLogin)
        if parent!=False:
            time.sleep(1)
            driver.back()
            time.sleep(1)
            driver.back()
            writeOnlyParent(parent[0], parent[1], parent[2], parent[3])
        else:
            driver.find_element(By.XPATH, "//input[@name='change']").click()
            fullName = driver.find_element(By.XPATH, "//div[@class= 'owner people_page' ]/div//h2").text
            tempPass = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[2]").text
            writeOnlyParent(fullName, studLogin, tempPass,user)
            time.sleep(1)
            driver.back()
            time.sleep(1)
            driver.back()
            time.sleep(1)
            driver.back()

def goToStudent(driver,parent=False):
    driver.find_element(By.XPATH,"//table[contains(@class,'people')]/tbody/tr[1]/td[@class = 'tdButtons']//a[@title='Редактировать']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']")))

    driver.find_element(By.XPATH,"//div[@class ='tabs']//li//a[@id='TabPassword']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='change']")))
    driver.find_element(By.XPATH,"//input[@name='change']").click()



    fullName = driver.find_element(By.XPATH,"//div[@class= 'owner people_page' ]/div//h2").text
    studLogin = driver.find_element(By.XPATH,"//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text
    tempPass = driver.find_element(By.XPATH,"//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[2]").text
    driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabReview']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='panel blue2'][2]//dl[@class='info big']/dd[1]//a")))
    classNumber = driver.find_element(By.XPATH,"//div[@class='panel blue2'][2]//dl[@class='info big']/dd[1]//a").text


    #Parents part
    driver.find_element(By.XPATH,"//div[@class ='tabs']//li//a[@id='TabParents']").click()

    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//li/a[@id='buttonAddRelative']")))
    parents = None
    if parent:
        try:
            driver.find_element(By.XPATH,"//div[@class='emptyData']")
        except:
            logins = getOnlyLoginParent(driver)
            status = False
            if logins !=0:
                status = checkParentsExist(logins)
            if status:
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                driver.back()
                a = []
                a.append(status)
                create_write_excel(fullName, studLogin, tempPass, classNumber, a)
            else:
                parents = goToParent(driver)
                create_write_excel(fullName, studLogin, tempPass, classNumber, parents)
    else:
        create_write_excel(fullName, studLogin, tempPass, classNumber, parents)

def goToParent(driver):
    parents = []
    driver.find_element(By.XPATH, f"//table[contains(@class,'people')]/tbody/tr[2]/td[@class='tdButtons']//li[@class='iE']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']")))
    driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='change']")))
    driver.find_element(By.XPATH, "//input[@name='change']").click()
    time.sleep(1)

    fullParentName = driver.find_element(By.XPATH, "//div[@class= 'owner people_page']/div//h2").text
    parentLogin = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text
    parentPass = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[2]").text
    daughterList = []
    daughterList.append(fullParentName)
    daughterList.append(parentLogin)
    daughterList.append(parentPass)
    daughterList.append('Родитель')
    parents.append(daughterList)
    driver.back()
    time.sleep(1)
    driver.back()
    time.sleep(1)
    driver.back()
    time.sleep(1)
    driver.back()
    time.sleep(1)
    driver.back()
    time.sleep(1)
    driver.back()
    time.sleep(1)
    driver.back()
    time.sleep(1)
    driver.back()
    return parents

def create_write_excel(studName,studLogin,studPass,classNumber = -1,parents:list=None):
    wb = None
    if os.path.isfile(f"{date.today()}_Отчет о выгрузке{nameExcel}.xlsx"):
        wb = openpyxl.load_workbook(f"{date.today()}_Отчет о выгрузке{nameExcel}.xlsx")

        sheet = wb.worksheets[0]
        last_empty_row = sheet.max_row + 1




        sheet.cell(row=last_empty_row, column=1).value = 'Ученик'
        sheet.cell(row=last_empty_row, column=2).value = studName
        sheet.cell(row=last_empty_row, column=3).value = studLogin
        sheet.cell(row=last_empty_row, column=4).value = studPass

        if classNumber != -1:
            sheet.cell(row=last_empty_row, column=5).value = classNumber
        if parents:
            parentName = parents[0][1]
            foundParent = 0
            for row in sheet.iter_rows(min_col=3, max_col=3):
                for cell in row:
                    if str(cell.value) == str(parentName):
                        foundParent = cell

            if foundParent != 0:
                row = foundParent.row
                sheet.cell(row=last_empty_row, column=6).value = sheet[row][6].value
                sheet.cell(row=last_empty_row, column=7).value = sheet[row][7].value
                sheet.cell(row=last_empty_row, column=8).value = sheet[row][8].value

            else:

                for i in range(len(parents)):
                    sheet.cell(row=last_empty_row, column=6).value = parents[i][0]
                    sheet.cell(row=last_empty_row, column=7).value = parents[i][1]
                    sheet.cell(row=last_empty_row, column=8).value = parents[i][2]

        wb.save(f'{date.today()}_Отчет о выгрузке{nameExcel}.xlsx')

    else:
        pass

def gotoDirectParent(driver):
    driver.find_element(By.XPATH,
                        "//table[contains(@class,'people')]/tbody//td[@class='tdButtons']//li[@class='iE']").click()
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']")))

    driver.find_element(By.XPATH, "//div[@class ='tabs']//li//a[@id='TabPassword']").click()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='change']")))

    studLogin = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[1]").text


    parent = checkParentsExist(studLogin)
    if parent:
        writeOnlyParent(parent[0],parent[1],parent[2],parent[3])
    else:
        driver.find_element(By.XPATH, "//input[@name='change']").click()
        fullName = driver.find_element(By.XPATH, "//div[@class= 'owner people_page' ]/div//h2").text
        tempPass = driver.find_element(By.XPATH, "//div[@class= 'panel']/div//dl[contains(@class,'info')]/dd[2]").text

        writeOnlyParent(fullName,studLogin,tempPass,"Родитель")

def writeOnlyParent(parentName,parentLogin,parentPass,user):
    if os.path.isfile(f"{date.today()}_Отчет о выгрузке{nameExcel}.xlsx"):
        wb = openpyxl.load_workbook(f"{date.today()}_Отчет о выгрузке{nameExcel}.xlsx")

        sheet = wb.worksheets[0]

        last_empty_row = sheet.max_row + 1
        sheet.cell(row=last_empty_row, column=1).value = user
        sheet.cell(row=last_empty_row, column=2).value = parentName
        sheet.cell(row=last_empty_row, column=3).value = parentLogin
        sheet.cell(row=last_empty_row, column=4).value = parentPass

        wb.save(f'{date.today()}_Отчет о выгрузке{nameExcel}.xlsx')
    else:
        pass

def createExcel(name):

    if os.path.isfile(f"{date.today()}_Отчет о выгрузке{name}.xlsx"):
        wb = openpyxl.load_workbook(f"{date.today()}_Отчет о выгрузке{name}.xlsx")
    else:
        wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]
    sheet.cell(row=1, column=1).value = "Роль"
    sheet.cell(row=1, column=2).value = "ФИО пользователя"
    sheet.cell(row=1, column=3).value = "Логин"
    sheet.cell(row=1, column=4).value = "Временный пароль"
    sheet.cell(row=1, column=5).value = "Класс"
    sheet.cell(row=1, column=6).value = "ФИО законного представителя"
    sheet.cell(row=1, column=7).value = "Логин"
    sheet.cell(row=1, column=8).value = "Временный пароль"
    wb.save(f'{date.today()}_Отчет о выгрузке{name}.xlsx')



def checkParentsExist(parentLogin):
    if os.path.isfile(f"{date.today()}_Отчет о выгрузке{nameExcel}.xlsx"):
        wb = openpyxl.load_workbook(f"{date.today()}_Отчет о выгрузке{nameExcel}.xlsx")
    else:
        wb = openpyxl.Workbook()

    sheet = wb.worksheets[0]
    foundParent = 0
    for row in sheet.iter_rows(min_col=7, max_col=7):
        for cell in row:
            if str(cell.value) == str(parentLogin):
                print(cell)
                foundParent = cell

    if foundParent != 0:
        parent = []
        row = foundParent.row
        nameParent = sheet[row][5].value
        loginParent = sheet[row][6].value
        passParent = sheet[row][7].value
        parent.append(nameParent)
        parent.append(loginParent)
        parent.append(passParent)
        return parent
    return False


    # create_write_excel("new2",'studLogin','pass',2,[['parentlogin','new1','sdwasdw']])
##1-ученик
# 2-родитель
# 3-ученик и родитель
# 4 teacher
# 5 сотрудники
if __name__=='__main__':
    # threading.Timer(1.25, webbrowser.open('http://127.0.0.1:8989/')).start()
    # app.run(port=8989, debug=True)
    main(roles=1,name = 'Иванов')


